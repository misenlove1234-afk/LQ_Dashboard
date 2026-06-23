"""
엑셀 회의록 제거 프로젝트 — 기준정보 데이터 레이어
DB 테이블 초기화, 조회, 수정 함수 모음
"""

import io
import logging
import traceback
import datetime
import pandas as pd
import streamlit as st
from utils.db import run_query, execute_query, get_connection

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════
# 초기 기준 데이터 (상수 — Streamlit 내에서 수정 가능)
# ═══════════════════════════════════════════════════════════════

# ── 30 STG 기준파일 ──────────────────────────────────────────
# vessel_type, block_no, 관철, 덕트, 전장, 도장, 배선
_INIT_30STG = [
    ("LNG","M51NS", 0,0,0,0,0), ("LNG","M51NP", 0,0,0,0,0),
    ("LNG","M900S", 2,0,2,0,0), ("LNG","M900P", 2,0,2,0,0),
    ("LNG","M610C", 4,4,5,2,3),
    ("LNG","M510S", 4,5,6,2,6), ("LNG","M510P", 4,5,6,2,6),
    ("LNG","M410S", 4,4,6,3,2), ("LNG","M410P", 4,4,6,3,3),
    ("LNG","M31NS", 1,0,0,0,0), ("LNG","M31NP", 1,0,0,0,0),
    ("LNG","M31OS", 4,4,6,2,4), ("LNG","M31OP", 4,4,6,2,4),
    ("LNG","M220S", 4,5,7,2,3), ("LNG","M220P", 4,5,7,2,3),
    ("LNG","M210S", 4,4,4,2,5), ("LNG","M210P", 4,4,4,2,5),
    ("LNG","M120S", 4,5,5,2,4), ("LNG","M120P", 4,5,5,2,4),
    ("LNG","M110S", 4,5,7,2,4), ("LNG","M110P", 4,5,7,2,4),
    ("CONT","M940S", 1,1,0,1,1), ("CONT","M930P", 1,1,0,1,1),
    ("CONT","M920S", 3,4,2,3,4), ("CONT","M920P", 3,4,2,3,4),
    ("CONT","M910S", 3,5,5,2,3), ("CONT","M910P", 3,5,5,2,3),
    ("CONT","M900S", 3,0,2,3,6), ("CONT","M900P", 3,0,2,3,6),
    ("CONT","M710S", 5,5,4,3,4), ("CONT","M710P", 5,5,4,3,4),
    ("CONT","M610S", 5,4,5,3,4), ("CONT","M610P", 5,4,5,3,4),
    ("CONT","M510S", 4,5,5,3,4), ("CONT","M510P", 4,5,5,3,4),
    ("CONT","M410S", 3,4,4,3,4), ("CONT","M410P", 3,4,4,3,4),
    ("CONT","M320S", 3,0,1,1,0), ("CONT","M320P", 3,0,1,1,0),
    ("CONT","M310S", 4,5,5,3,5), ("CONT","M310P", 4,5,5,3,5),
    ("CONT","M220S", 4,4,2,3,0), ("CONT","M220P", 4,4,2,3,0),
    ("CONT","M210S", 5,5,6,3,5), ("CONT","M210P", 5,5,6,3,5),
    ("CONT","M120S", 2,4,2,1,0), ("CONT","M120P", 2,4,2,1,0),
    ("CONT","M110S", 5,5,6,3,3), ("CONT","M110P", 5,5,6,3,3),
]

# ── 50 STG 기준파일 ──────────────────────────────────────────
# 열 순서: vessel_type, deck,
#          선각취부, 선각용접, 선각FLOOR곡직, 선각WALL곡직,   ← 선각 4개
#          목의화기1차, 도장PB, 도장TU, 배선, 보온,
#          목의화기2차, 피복, BP검사, 판넬설치, 결선, 가구, 장판, 스커트
_INIT_50STG = [
    # 취부,용접,FLOOR,WALL 순 (LNG: 9,6,4,8 / CONT: 5,5,4,4)
    ("LNG","UPP-DK", 9,6,4,8,  0,2,3, 0, 0, 0,0,0, 0, 0,0,0,0),
    ("LNG","A-DK",   9,6,4,8,  7,2,3, 5, 4, 2,1,1, 4, 5,1,1,1),
    ("LNG","B-DK",   9,6,4,8,  6,2,3, 5, 4, 3,1,1, 4, 5,1,1,1),
    ("LNG","C-DK",   9,6,4,8,  6,2,3, 5, 4, 3,1,1, 4, 5,1,1,1),
    ("LNG","D-DK",   9,6,4,8,  6,2,3,15, 8, 3,1,1, 4, 5,1,1,1),
    ("LNG","NAV-DK", 9,6,4,8,  4,2,3,10, 5, 0,1,1, 4, 5,1,1,1),
    ("CONT","UPP-DK",5,5,4,4,  0,2,3, 0, 0, 0,0,0, 0, 0,0,0,0),
    ("CONT","A-DK",  5,5,4,4,  5,2,3, 4, 3, 2,1,1, 3, 3,2,1,1),
    ("CONT","B-DK",  5,5,4,4,  5,2,3, 4, 3, 3,1,1, 3, 3,2,1,1),
    ("CONT","C-DK",  5,5,4,4,  5,2,3, 4, 3, 2,1,1, 3, 3,2,1,1),
    ("CONT","D-DK",  5,5,4,4,  5,2,3, 4, 3, 2,1,1, 3, 3,2,1,1),
    ("CONT","E-DK",  5,5,4,4,  5,2,3, 4, 3, 3,1,1, 3, 3,2,1,1),
    ("CONT","F-DK",  5,5,4,4,  5,2,3,15, 5, 3,1,1, 3, 3,2,1,1),
    ("CONT","NAV-DK",5,5,4,4,  6,2,3,10, 5, 0,1,1, 3, 3,2,1,1),
]

# ── In/Outside / STAIRWAY / UPP-DK 기준파일 ─────────────────
# 열 순서: area,
#   [In/Outside·Elevator] 트렁크배선,윈도우설치,윈도우검사,내부배관검사,외부배관검사,
#                          외판도장,외판족장철거,FLOOR도장,목의장비탑재,목의장비설치,
#                          탑재준비,탑재사열,인양,탑재,
#   [STAIRWAY·UPP-DK]     도장PB,도장TU,보온,판넬설치,결선,장판,스커트,배선,족장철거
_INIT_INOUT = [
    ("In/Outside", 5,3,1,1,1,40,3,4,0,0, 2,1,1,1, 0,0,0,0,0,0,0,0,0),
    ("Elevator",   3,0,0,0,0, 0,0,0,0,0, 0,0,0,0, 0,0,0,0,0,0,0,0,0),
    ("STAIRWAY",   0,0,0,0,0, 0,0,0,0,0, 0,0,0,0, 0,0,0,0,0,0,0,0,0),
    ("UPP-DK",     0,0,0,0,0, 0,0,0,0,0, 0,0,0,0, 0,0,0,0,0,0,0,0,0),
]

# In/Outside 계열 공통 작업 컬럼 목록 (seed·query·save에 재사용)
_IO_WORK_COLS = [
    '트렁크배선','윈도우설치','윈도우검사','내부배관검사','외부배관검사',
    '외판도장','외판족장철거','FLOOR도장','목의장비탑재','목의장비설치',
    '탑재준비','탑재사열','인양','탑재',
    '도장PB','도장TU','보온','판넬설치','결선','장판','스커트','배선','족장철거',
]

# ── 공종 실행 순서 ────────────────────────────────────────────
# stg, seq_no, work_code, work_name,
# duration_type('anchor'|'fixed'|'variable'), fixed_days,
# exception_blocks(쉼표 구분), cal_type('all'|'weekday'), ref_col
_INIT_SEQUENCE = [
    # ─── 30 STG ─────────────────────────────────────────────────
    ("30", 1, "blk_in",       "블럭입고",       "anchor",   0, "",                            "all",     ""),
    ("30", 2, "족장설치",      "족장설치",       "fixed",    1, "M31NS,M31NP,M51NS,M51NP",    "all",     ""),
    ("30", 3, "관철",          "관철",           "variable", 0, "",                            "all",     "관철"),
    ("30", 4, "덕트",          "덕트",           "variable", 0, "",                            "all",     "덕트"),
    ("30", 5, "전장",          "전장",           "variable", 0, "",                            "all",     "전장"),
    ("30", 6, "도장",          "선도장",         "variable", 0, "",                            "all",     "도장"),
    ("30", 7, "배선",          "배선",           "variable", 0, "",                            "all",     "배선"),
    ("30", 8, "뒤집기",        "뒤집기",         "fixed",    1, "",                            "all",     ""),
    ("30", 9, "blk_out",       "블럭탑재",       "anchor",   0, "",                            "all",     ""),
    # ─── 50 STG ─────────────────────────────────────────────────
    ("50",  1, "블럭탑재",      "블럭탑재",       "anchor",   0, "", "all",     ""),
    ("50",  2, "선각취부",      "선각취부",       "variable", 0, "", "all",     "선각취부"),
    ("50",  3, "선각용접",      "선각용접",       "variable", 0, "", "all",     "선각용접"),
    ("50",  4, "선각FLOOR곡직", "선각FLOOR곡직",  "variable", 0, "", "all",     "선각FLOOR곡직"),
    ("50",  5, "선각WALL곡직",  "선각WALL곡직",   "variable", 0, "", "all",     "선각WALL곡직"),
    ("50",  6, "선각검사",      "선각검사",       "anchor",   0, "", "all",     ""),
    ("50",  7, "목의화기1차",   "목의화기1차",    "variable", 0, "", "all",     "목의화기1차"),
    ("50",  8, "도장PB",        "도장PB",         "variable", 0, "", "all",     "도장PB"),
    ("50",  9, "도장TU",        "도장TU",         "variable", 0, "", "all",     "도장TU"),
    ("50", 10, "배선",          "배선",           "variable", 0, "", "all",     "배선"),
    ("50", 11, "보온",          "보온",           "variable", 0, "", "all",     "보온"),
    ("50", 12, "목의화기2차",   "목의화기2차",    "variable", 0, "", "all",     "목의화기2차"),
    ("50", 13, "피복",          "피복",           "variable", 0, "", "all",     "피복"),
    ("50", 14, "BP검사",        "B/P검사",        "variable", 0, "", "weekday", "BP검사"),
    ("50", 15, "판넬설치",      "판넬설치",       "variable", 0, "", "all",     "판넬설치"),
    ("50", 16, "결선",          "결선",           "variable", 0, "", "all",     "결선"),
    ("50", 17, "가구",          "가구",           "variable", 0, "", "all",     "가구"),
    ("50", 18, "장판",          "장판",           "variable", 0, "", "all",     "장판"),
    ("50", 19, "스커트",        "스커트",         "variable", 0, "", "all",     "스커트"),
]

# ── 데크 순서 (UPP=하부=1, NAV=상부=최대) ─────────────────────
# vessel_type, deck, order_no
_INIT_DECK_ORDER = [
    ("LNG",  "UPP-DK", 1), ("LNG",  "A-DK",   2), ("LNG",  "B-DK",  3),
    ("LNG",  "C-DK",   4), ("LNG",  "D-DK",   5), ("LNG",  "NAV-DK",6),
    ("CONT", "UPP-DK", 1), ("CONT", "A-DK",   2), ("CONT", "B-DK",  3),
    ("CONT", "C-DK",   4), ("CONT", "D-DK",   5), ("CONT", "E-DK",  6),
    ("CONT", "F-DK",   7), ("CONT", "NAV-DK", 8),
]

# ── 블럭 → 데크 매핑 (30STG 블럭이 속한 50STG 데크) ──────────
# vessel_type, block_no, deck
_INIT_BLOCK_DECK_MAP = [
    # LNG (21 블럭)
    ("LNG","M51NS","NAV-DK"), ("LNG","M51NP","NAV-DK"),
    ("LNG","M900S","NAV-DK"), ("LNG","M900P","NAV-DK"),
    ("LNG","M610C","D-DK"),
    ("LNG","M510S","D-DK"),   ("LNG","M510P","D-DK"),
    ("LNG","M410S","C-DK"),   ("LNG","M410P","C-DK"),
    ("LNG","M31NS","B-DK"),   ("LNG","M31NP","B-DK"),
    ("LNG","M31OS","B-DK"),   ("LNG","M31OP","B-DK"),
    ("LNG","M220S","A-DK"),   ("LNG","M220P","A-DK"),
    ("LNG","M210S","A-DK"),   ("LNG","M210P","A-DK"),
    ("LNG","M120S","UPP-DK"), ("LNG","M120P","UPP-DK"),
    ("LNG","M110S","UPP-DK"), ("LNG","M110P","UPP-DK"),
    # CONT (28 블럭)
    ("CONT","M940S","NAV-DK"), ("CONT","M930P","NAV-DK"),
    ("CONT","M920S","NAV-DK"), ("CONT","M920P","NAV-DK"),
    ("CONT","M910S","NAV-DK"), ("CONT","M910P","NAV-DK"),
    ("CONT","M900S","NAV-DK"), ("CONT","M900P","NAV-DK"),
    ("CONT","M710S","F-DK"),   ("CONT","M710P","F-DK"),
    ("CONT","M610S","E-DK"),   ("CONT","M610P","E-DK"),
    ("CONT","M510S","D-DK"),   ("CONT","M510P","D-DK"),
    ("CONT","M410S","C-DK"),   ("CONT","M410P","C-DK"),
    ("CONT","M320S","B-DK"),   ("CONT","M320P","B-DK"),
    ("CONT","M310S","B-DK"),   ("CONT","M310P","B-DK"),
    ("CONT","M220S","A-DK"),   ("CONT","M220P","A-DK"),
    ("CONT","M210S","A-DK"),   ("CONT","M210P","A-DK"),
    ("CONT","M120S","A-DK"),   ("CONT","M120P","A-DK"),
    ("CONT","M110S","UPP-DK"), ("CONT","M110P","UPP-DK"),
]

# ── 로직 규칙 ────────────────────────────────────────────────
_INIT_RULES = [
    # rule_id, vessel_type, rule_name,
    # pre_area, pre_work, pre_event, offset_days,
    # suc_area, suc_work, is_active
    (1,"ALL","화기→도장 리드",
     "upper_deck","목의화기1차","END",-2,
     "lower_deck","도장PB",1),
    (2,"ALL","트렁크→배선",
     "in_outside","트렁크배선","END",0,
     "all_deck","배선",1),
    (3,"ALL","선각검사→의장",
     "deck","선각검사","END",0,
     "deck","목의화기1차",1),
    (4,"ALL","계단식 배치",
     "lower_deck","same","END",0,
     "upper_deck","same",1),
    (5,"LNG","D-DK WALL→트렁크(LNG)",
     "D-DK","선각WALL곡직","END",0,
     "in_outside","트렁크배선",1),
    (6,"CONT","F-DK WALL→트렁크(CONT)",
     "F-DK","선각WALL곡직","END",0,
     "in_outside","트렁크배선",1),
]


# ═══════════════════════════════════════════════════════════════
# DDL
# ═══════════════════════════════════════════════════════════════

_DDL_30STG = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_ref_30stg' AND xtype='U')
CREATE TABLE lq_meet_ref_30stg (
    id          INT IDENTITY(1,1) PRIMARY KEY,
    vessel_type NVARCHAR(10)  NOT NULL,
    block_no    NVARCHAR(20)  NOT NULL,
    관철         INT DEFAULT 0,
    덕트         INT DEFAULT 0,
    전장         INT DEFAULT 0,
    도장         INT DEFAULT 0,
    배선         INT DEFAULT 0,
    updated_at  DATETIME DEFAULT GETDATE(),
    CONSTRAINT uq_30stg UNIQUE (vessel_type, block_no)
)
"""

# 선각 4개 컬럼 포함한 전체 DDL
_DDL_50STG = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_ref_50stg' AND xtype='U')
CREATE TABLE lq_meet_ref_50stg (
    id            INT IDENTITY(1,1) PRIMARY KEY,
    vessel_type   NVARCHAR(10)  NOT NULL,
    deck          NVARCHAR(20)  NOT NULL,
    선각취부       INT DEFAULT 0,
    선각용접       INT DEFAULT 0,
    선각FLOOR곡직  INT DEFAULT 0,
    선각WALL곡직   INT DEFAULT 0,
    목의화기1차    INT DEFAULT 0,
    도장PB        INT DEFAULT 0,
    도장TU        INT DEFAULT 0,
    배선          INT DEFAULT 0,
    보온          INT DEFAULT 0,
    목의화기2차    INT DEFAULT 0,
    피복          INT DEFAULT 0,
    BP검사        INT DEFAULT 0,
    판넬설치       INT DEFAULT 0,
    결선          INT DEFAULT 0,
    가구          INT DEFAULT 0,
    장판          INT DEFAULT 0,
    스커트         INT DEFAULT 0,
    updated_at    DATETIME DEFAULT GETDATE(),
    CONSTRAINT uq_50stg UNIQUE (vessel_type, deck)
)
"""

# 기존 테이블에 선각 컬럼이 없을 때 추가 (마이그레이션)
_ALTER_50STG_COLS = [
    "선각취부",
    "선각용접",
    "선각FLOOR곡직",
    "선각WALL곡직",
]

# 기존 lq_meet_anchor_50stg에 신규 컬럼 마이그레이션
_ALTER_ANCHOR_50STG_COLS = [
    ("wall_straight_date",  "DATE"),
    ("sunggak_attach_end",  "DATE"),
    ("sunggak_weld_end",    "DATE"),
    ("floor_straight_date", "DATE"),
    ("inspection_date",     "DATE"),
]

_DDL_INOUT = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_ref_inout' AND xtype='U')
CREATE TABLE lq_meet_ref_inout (
    id           INT IDENTITY(1,1) PRIMARY KEY,
    area         NVARCHAR(30) NOT NULL UNIQUE,
    트렁크배선    INT DEFAULT 0,
    윈도우설치    INT DEFAULT 0,
    윈도우검사    INT DEFAULT 0,
    내부배관검사  INT DEFAULT 0,
    외부배관검사  INT DEFAULT 0,
    외판도장      INT DEFAULT 0,
    외판족장철거  INT DEFAULT 0,
    FLOOR도장     INT DEFAULT 0,
    목의장비탑재  INT DEFAULT 0,
    목의장비설치  INT DEFAULT 0,
    탑재준비      INT DEFAULT 0,
    탑재사열      INT DEFAULT 0,
    인양         INT DEFAULT 0,
    탑재         INT DEFAULT 0,
    도장PB        INT DEFAULT 0,
    도장TU        INT DEFAULT 0,
    보온          INT DEFAULT 0,
    판넬설치      INT DEFAULT 0,
    결선          INT DEFAULT 0,
    장판          INT DEFAULT 0,
    스커트        INT DEFAULT 0,
    배선          INT DEFAULT 0,
    족장철거      INT DEFAULT 0,
    updated_at   DATETIME DEFAULT GETDATE()
)
"""

# 기존 lq_meet_ref_inout에 신규 컬럼 마이그레이션
_ALTER_INOUT_COLS = [
    "목의장비탑재", "목의장비설치",
    "도장PB", "도장TU", "보온", "판넬설치", "결선", "장판", "스커트", "배선", "족장철거",
]

_DDL_SEQUENCE = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_ref_sequence' AND xtype='U')
CREATE TABLE lq_meet_ref_sequence (
    id               INT IDENTITY(1,1) PRIMARY KEY,
    stg              NVARCHAR(5)    NOT NULL,
    seq_no           INT            NOT NULL,
    work_code        NVARCHAR(30)   NOT NULL,
    work_name        NVARCHAR(50)   NOT NULL,
    duration_type    NVARCHAR(20)   NOT NULL,
    fixed_days       INT            DEFAULT 0,
    exception_blocks NVARCHAR(200)  DEFAULT '',
    cal_type         NVARCHAR(20)   DEFAULT 'all',
    ref_col          NVARCHAR(50)   DEFAULT '',
    is_active        INT            DEFAULT 1,
    updated_at       DATETIME       DEFAULT GETDATE(),
    CONSTRAINT uq_sequence UNIQUE (stg, seq_no)
)
"""

_DDL_DECK_ORDER = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_ref_deck_order' AND xtype='U')
CREATE TABLE lq_meet_ref_deck_order (
    id          INT IDENTITY(1,1) PRIMARY KEY,
    vessel_type NVARCHAR(10)  NOT NULL,
    deck        NVARCHAR(20)  NOT NULL,
    order_no    INT           NOT NULL,
    updated_at  DATETIME      DEFAULT GETDATE(),
    CONSTRAINT uq_deck_order UNIQUE (vessel_type, deck)
)
"""

_DDL_RULES = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_logic_rules' AND xtype='U')
CREATE TABLE lq_meet_logic_rules (
    rule_id      INT PRIMARY KEY,
    vessel_type  NVARCHAR(10)  NOT NULL,
    rule_name    NVARCHAR(100) NOT NULL,
    pre_area     NVARCHAR(50)  NOT NULL,
    pre_work     NVARCHAR(50)  NOT NULL,
    pre_event    NVARCHAR(10)  NOT NULL,
    offset_days  INT           NOT NULL DEFAULT 0,
    suc_area     NVARCHAR(50)  NOT NULL,
    suc_work     NVARCHAR(50)  NOT NULL,
    is_active    INT           NOT NULL DEFAULT 1,
    updated_at   DATETIME      DEFAULT GETDATE()
)
"""

_DDL_CALENDAR = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_calendar' AND xtype='U')
CREATE TABLE lq_meet_calendar (
    id         INT IDENTITY(1,1) PRIMARY KEY,
    cal_date   DATE         NOT NULL UNIQUE,
    reason     NVARCHAR(50) NOT NULL
)
"""

# ── 앵커 이벤트 테이블 ─────────────────────────────────────────

_DDL_VESSEL = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_vessel' AND xtype='U')
CREATE TABLE lq_meet_vessel (
    vessel_no        NVARCHAR(20)  NOT NULL PRIMARY KEY,
    vessel_type      NVARCHAR(10)  NOT NULL,
    거주구탑재예정일  DATE,
    비고              NVARCHAR(200),
    updated_at       DATETIME      DEFAULT GETDATE()
)
"""

_DDL_ANCHOR_30STG = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_anchor_30stg' AND xtype='U')
CREATE TABLE lq_meet_anchor_30stg (
    id           INT IDENTITY(1,1) PRIMARY KEY,
    vessel_no    NVARCHAR(20)  NOT NULL,
    block_no     NVARCHAR(20)  NOT NULL,
    blk_in_date  DATE,
    blk_out_date DATE,
    updated_at   DATETIME      DEFAULT GETDATE(),
    CONSTRAINT uq_anchor30 UNIQUE (vessel_no, block_no)
)
"""

_DDL_ANCHOR_50STG = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_anchor_50stg' AND xtype='U')
CREATE TABLE lq_meet_anchor_50stg (
    id                  INT IDENTITY(1,1) PRIMARY KEY,
    vessel_no           NVARCHAR(20)   NOT NULL,
    deck                NVARCHAR(20)   NOT NULL,
    mount_start_date    DATE,
    sunggak_attach_end  DATE,
    sunggak_weld_end    DATE,
    floor_straight_date DATE,
    wall_straight_date  DATE,
    inspection_date     DATE,
    updated_at          DATETIME       DEFAULT GETDATE(),
    CONSTRAINT uq_anchor50 UNIQUE (vessel_no, deck)
)
"""

_DDL_BLOCK_DECK_MAP = """
IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='lq_meet_ref_block_deck' AND xtype='U')
CREATE TABLE lq_meet_ref_block_deck (
    id          INT IDENTITY(1,1) PRIMARY KEY,
    vessel_type NVARCHAR(10)  NOT NULL,
    block_no    NVARCHAR(20)  NOT NULL,
    deck        NVARCHAR(20)  NOT NULL,
    updated_at  DATETIME      DEFAULT GETDATE(),
    CONSTRAINT uq_block_deck UNIQUE (vessel_type, block_no)
)
"""


# ═══════════════════════════════════════════════════════════════
# 테이블 초기화
# ═══════════════════════════════════════════════════════════════

def init_tables() -> bool:
    """DB 테이블이 없으면 생성하고 초기 데이터를 INSERT"""
    try:
        conn = get_connection()
        cur = conn.cursor()

        for ddl in [_DDL_30STG, _DDL_50STG, _DDL_INOUT,
                    _DDL_SEQUENCE, _DDL_DECK_ORDER,
                    _DDL_RULES, _DDL_CALENDAR,
                    _DDL_VESSEL, _DDL_ANCHOR_30STG, _DDL_ANCHOR_50STG,
                    _DDL_BLOCK_DECK_MAP]:
            cur.execute(ddl)
        conn.commit()

        # 기존 lq_meet_ref_50stg에 선각 컬럼이 없으면 추가 (마이그레이션)
        for col in _ALTER_50STG_COLS:
            cur.execute(f"""
                IF NOT EXISTS (
                    SELECT * FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME='lq_meet_ref_50stg' AND COLUMN_NAME=N'{col}'
                )
                ALTER TABLE lq_meet_ref_50stg ADD [{col}] INT DEFAULT 0
            """)
        # 기존 lq_meet_anchor_50stg에 wall_straight_date 컬럼 마이그레이션
        for col, col_type in _ALTER_ANCHOR_50STG_COLS:
            cur.execute(f"""
                IF NOT EXISTS (
                    SELECT * FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME='lq_meet_anchor_50stg' AND COLUMN_NAME=N'{col}'
                )
                ALTER TABLE lq_meet_anchor_50stg ADD [{col}] {col_type}
            """)
        # 기존 lq_meet_ref_inout에 신규 컬럼 마이그레이션 (STAIRWAY/UPP-DK 추가)
        for col in _ALTER_INOUT_COLS:
            cur.execute(f"""
                IF NOT EXISTS (
                    SELECT * FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME='lq_meet_ref_inout' AND COLUMN_NAME=N'{col}'
                )
                ALTER TABLE lq_meet_ref_inout ADD [{col}] INT DEFAULT 0
            """)
        conn.commit()

        # 선각 소요일 기준값 동기화 — _INIT_50STG 값으로 기존 행 강제 업데이트
        for row in _INIT_50STG:
            vt, dk = row[0], row[1]
            at, wd, fl, wl = row[2], row[3], row[4], row[5]
            cur.execute("""
                UPDATE lq_meet_ref_50stg
                SET 선각취부=?,선각용접=?,[선각FLOOR곡직]=?,[선각WALL곡직]=?
                WHERE vessel_type=? AND deck=?
                  AND (선각취부=0 OR 선각취부 IS NULL)
            """, (at, wd, fl, wl, vt, dk))
        conn.commit()

        _seed_30stg(cur)
        _seed_50stg(cur)
        _seed_inout(cur)
        _seed_sequence(cur)
        _seed_deck_order(cur)
        _seed_rules(cur)
        _seed_block_deck_map(cur)
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.error("테이블 초기화 오류: %s\n%s", e, traceback.format_exc())
        return False


# ── 시드 함수들 ───────────────────────────────────────────────

def _seed_30stg(cur):
    for row in _INIT_30STG:
        cur.execute("""
            IF NOT EXISTS (SELECT 1 FROM lq_meet_ref_30stg WHERE vessel_type=? AND block_no=?)
            INSERT INTO lq_meet_ref_30stg (vessel_type,block_no,관철,덕트,전장,도장,배선)
            VALUES (?,?,?,?,?,?,?)
        """, (row[0], row[1], row[0], row[1], row[2], row[3], row[4], row[5], row[6]))


def _seed_50stg(cur):
    for row in _INIT_50STG:
        cur.execute("""
            IF NOT EXISTS (SELECT 1 FROM lq_meet_ref_50stg WHERE vessel_type=? AND deck=?)
            INSERT INTO lq_meet_ref_50stg
              (vessel_type,deck,
               선각취부,선각용접,[선각FLOOR곡직],[선각WALL곡직],
               목의화기1차,도장PB,도장TU,배선,보온,
               목의화기2차,피복,BP검사,판넬설치,결선,가구,장판,스커트)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (row[0],row[1],
              row[0],row[1],row[2],row[3],row[4],row[5],
              row[6],row[7],row[8],row[9],row[10],
              row[11],row[12],row[13],row[14],row[15],row[16],row[17],row[18]))


def _seed_inout(cur):
    col_str = ','.join(f'[{c}]' for c in _IO_WORK_COLS)
    ph = ','.join(['?'] * len(_IO_WORK_COLS))
    for row in _INIT_INOUT:
        cur.execute(f"""
            IF NOT EXISTS (SELECT 1 FROM lq_meet_ref_inout WHERE area=?)
            INSERT INTO lq_meet_ref_inout (area,{col_str})
            VALUES (?,{ph})
        """, (row[0], row[0], *row[1:]))


def _seed_sequence(cur):
    for row in _INIT_SEQUENCE:
        cur.execute("""
            IF NOT EXISTS (SELECT 1 FROM lq_meet_ref_sequence WHERE stg=? AND seq_no=?)
            INSERT INTO lq_meet_ref_sequence
              (stg,seq_no,work_code,work_name,duration_type,fixed_days,
               exception_blocks,cal_type,ref_col,is_active)
            VALUES (?,?,?,?,?,?,?,?,?,1)
        """, (row[0],row[1],
              row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8]))


def _seed_deck_order(cur):
    for row in _INIT_DECK_ORDER:
        cur.execute("""
            IF NOT EXISTS (SELECT 1 FROM lq_meet_ref_deck_order WHERE vessel_type=? AND deck=?)
            INSERT INTO lq_meet_ref_deck_order (vessel_type,deck,order_no)
            VALUES (?,?,?)
        """, (row[0],row[1], row[0],row[1],row[2]))


def _seed_rules(cur):
    for row in _INIT_RULES:
        cur.execute("""
            IF NOT EXISTS (SELECT 1 FROM lq_meet_logic_rules WHERE rule_id=?)
            INSERT INTO lq_meet_logic_rules
              (rule_id,vessel_type,rule_name,pre_area,pre_work,pre_event,
               offset_days,suc_area,suc_work,is_active)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (row[0], row[0],row[1],row[2],row[3],row[4],row[5],
              row[6],row[7],row[8],row[9]))


def _seed_block_deck_map(cur):
    for row in _INIT_BLOCK_DECK_MAP:
        cur.execute("""
            IF NOT EXISTS (SELECT 1 FROM lq_meet_ref_block_deck WHERE vessel_type=? AND block_no=?)
            INSERT INTO lq_meet_ref_block_deck (vessel_type,block_no,deck)
            VALUES (?,?,?)
        """, (row[0], row[1], row[0], row[1], row[2]))


# ═══════════════════════════════════════════════════════════════
# 조회 함수
# ═══════════════════════════════════════════════════════════════

@st.cache_data(ttl=300)
def get_ref_30stg(vessel_type: str = None) -> pd.DataFrame:
    sql = "SELECT vessel_type,block_no,관철,덕트,전장,도장,배선 FROM lq_meet_ref_30stg"
    if vessel_type:
        sql += " WHERE vessel_type=?"
        return run_query(sql, params=(vessel_type,))
    return run_query(sql)


@st.cache_data(ttl=300)
def get_ref_50stg(vessel_type: str = None) -> pd.DataFrame:
    sql = """SELECT vessel_type,deck,
                    선각취부,선각용접,[선각FLOOR곡직],[선각WALL곡직],
                    목의화기1차,도장PB,도장TU,배선,보온,
                    목의화기2차,피복,BP검사,판넬설치,결선,가구,장판,스커트
             FROM lq_meet_ref_50stg"""
    if vessel_type:
        sql += " WHERE vessel_type=?"
        return run_query(sql, params=(vessel_type,))
    return run_query(sql)


@st.cache_data(ttl=300)
def get_ref_inout() -> pd.DataFrame:
    cols = ','.join(f'[{c}]' for c in _IO_WORK_COLS)
    return run_query(f"SELECT area,{cols} FROM lq_meet_ref_inout")


@st.cache_data(ttl=300)
def get_sequence(stg: str = None) -> pd.DataFrame:
    sql = """SELECT stg,seq_no,work_code,work_name,duration_type,fixed_days,
                    exception_blocks,cal_type,ref_col,is_active
             FROM lq_meet_ref_sequence"""
    if stg:
        sql += " WHERE stg=?"
        return run_query(sql + " ORDER BY seq_no", params=(stg,))
    return run_query(sql + " ORDER BY stg,seq_no")


@st.cache_data(ttl=300)
def get_deck_order(vessel_type: str = None) -> pd.DataFrame:
    sql = "SELECT vessel_type,deck,order_no FROM lq_meet_ref_deck_order"
    if vessel_type:
        sql += " WHERE vessel_type=?"
        return run_query(sql + " ORDER BY order_no", params=(vessel_type,))
    return run_query(sql + " ORDER BY vessel_type,order_no")


@st.cache_data(ttl=300)
def get_logic_rules() -> pd.DataFrame:
    return run_query("""SELECT rule_id,vessel_type,rule_name,pre_area,pre_work,
                               pre_event,offset_days,suc_area,suc_work,is_active
                        FROM lq_meet_logic_rules ORDER BY rule_id""")


@st.cache_data(ttl=300)
def get_calendar() -> pd.DataFrame:
    return run_query("""SELECT cal_date,reason FROM lq_meet_calendar
                        ORDER BY cal_date""")


@st.cache_data(ttl=300)
def get_block_deck_map(vessel_type: str = None) -> pd.DataFrame:
    sql = "SELECT vessel_type,block_no,deck FROM lq_meet_ref_block_deck"
    if vessel_type:
        sql += " WHERE vessel_type=?"
        return run_query(sql + " ORDER BY block_no", params=(vessel_type,))
    return run_query(sql + " ORDER BY vessel_type,block_no")


def get_block_to_deck_dict(vessel_type: str) -> dict:
    """블럭번호 → 데크 딕셔너리 반환 (계산 엔진용, 캐시 없음)"""
    df = get_block_deck_map(vessel_type)
    if df.empty:
        # DB 없을 때 상수에서 폴백
        rows = [r for r in _INIT_BLOCK_DECK_MAP if r[0] == vessel_type]
        return {r[1]: r[2] for r in rows}
    return dict(zip(df["block_no"], df["deck"]))


# ═══════════════════════════════════════════════════════════════
# 저장 함수
# ═══════════════════════════════════════════════════════════════

def save_ref_30stg(df: pd.DataFrame) -> bool:
    try:
        conn = get_connection()
        cur = conn.cursor()
        for _, row in df.iterrows():
            def _si(v):
                return int(v) if pd.notna(v) else 0
            cur.execute("""
                UPDATE lq_meet_ref_30stg
                SET 관철=?,덕트=?,전장=?,도장=?,배선=?,updated_at=GETDATE()
                WHERE vessel_type=? AND block_no=?
            """, (_si(row['관철']),_si(row['덕트']),_si(row['전장']),
                  _si(row['도장']),_si(row['배선']),
                  row['vessel_type'],row['block_no']))
        conn.commit()
        conn.close()
        get_ref_30stg.clear()
        return True
    except Exception as e:
        logger.error("30STG 저장 오류: %s\n%s", e, traceback.format_exc())
        return False


def save_ref_50stg(df: pd.DataFrame) -> bool:
    try:
        conn = get_connection()
        cur = conn.cursor()
        sunggak_cols = ['선각취부','선각용접','선각FLOOR곡직','선각WALL곡직']
        work_cols    = ['목의화기1차','도장PB','도장TU','배선','보온',
                        '목의화기2차','피복','BP검사','판넬설치','결선','가구','장판','스커트']
        all_cols = sunggak_cols + work_cols
        for _, row in df.iterrows():
            vals = [int(row[c]) if pd.notna(row.get(c)) else 0 for c in all_cols]
            set_clause = ",".join(f"[{c}]=?" for c in all_cols)
            cur.execute(f"""
                UPDATE lq_meet_ref_50stg
                SET {set_clause}, updated_at=GETDATE()
                WHERE vessel_type=? AND deck=?
            """, (*vals, row['vessel_type'], row['deck']))
        conn.commit()
        conn.close()
        get_ref_50stg.clear()
        return True
    except Exception as e:
        logger.error("50STG 저장 오류: %s\n%s", e, traceback.format_exc())
        return False


def save_ref_inout(df: pd.DataFrame) -> bool:
    try:
        conn = get_connection()
        cur = conn.cursor()
        cols = [c for c in _IO_WORK_COLS if c in df.columns]
        for _, row in df.iterrows():
            vals = [int(row[c]) for c in cols]
            set_clause = ','.join(f'[{c}]=?' for c in cols)
            cur.execute(f"""
                UPDATE lq_meet_ref_inout
                SET {set_clause}, updated_at=GETDATE()
                WHERE area=?
            """, (*vals, row['area']))
        conn.commit()
        conn.close()
        get_ref_inout.clear()
        return True
    except Exception as e:
        logger.error("InOut 저장 오류: %s\n%s", e, traceback.format_exc())
        return False


def save_sequence(df: pd.DataFrame) -> bool:
    try:
        conn = get_connection()
        cur = conn.cursor()
        for _, row in df.iterrows():
            cur.execute("""
                UPDATE lq_meet_ref_sequence
                SET work_name=?,duration_type=?,fixed_days=?,
                    exception_blocks=?,cal_type=?,ref_col=?,is_active=?,
                    updated_at=GETDATE()
                WHERE stg=? AND seq_no=?
            """, (row['work_name'],row['duration_type'],int(row['fixed_days']),
                  row['exception_blocks'],row['cal_type'],row['ref_col'],
                  int(row['is_active']), row['stg'],int(row['seq_no'])))
        conn.commit()
        conn.close()
        get_sequence.clear()
        return True
    except Exception as e:
        logger.error("시퀀스 저장 오류: %s\n%s", e, traceback.format_exc())
        return False


def save_logic_rules(df: pd.DataFrame) -> bool:
    try:
        conn = get_connection()
        cur = conn.cursor()
        for _, row in df.iterrows():
            cur.execute("""
                UPDATE lq_meet_logic_rules
                SET vessel_type=?,rule_name=?,pre_area=?,pre_work=?,pre_event=?,
                    offset_days=?,suc_area=?,suc_work=?,is_active=?,updated_at=GETDATE()
                WHERE rule_id=?
            """, (row['vessel_type'],row['rule_name'],row['pre_area'],row['pre_work'],
                  row['pre_event'],int(row['offset_days']),row['suc_area'],row['suc_work'],
                  int(row['is_active']),int(row['rule_id'])))
        conn.commit()
        conn.close()
        get_logic_rules.clear()
        return True
    except Exception as e:
        logger.error("로직 규칙 저장 오류: %s\n%s", e, traceback.format_exc())
        return False


def add_calendar_date(cal_date: str, reason: str) -> bool:
    try:
        execute_query("""
            IF NOT EXISTS (SELECT 1 FROM lq_meet_calendar WHERE cal_date=?)
            INSERT INTO lq_meet_calendar (cal_date,reason) VALUES (?,?)
        """, (cal_date, cal_date, reason))
        get_calendar.clear()
        return True
    except Exception as e:
        logger.error("캘린더 추가 오류: %s\n%s", e, traceback.format_exc())
        return False


def delete_calendar_date(cal_date: str) -> bool:
    try:
        execute_query("DELETE FROM lq_meet_calendar WHERE cal_date=?", (cal_date,))
        get_calendar.clear()
        return True
    except Exception as e:
        logger.error("캘린더 삭제 오류: %s\n%s", e, traceback.format_exc())
        return False


# ═══════════════════════════════════════════════════════════════
# 기본값 DataFrame (DB 없을 때 폴백용)
# ═══════════════════════════════════════════════════════════════

def default_30stg_df(vessel_type: str) -> pd.DataFrame:
    rows = [r for r in _INIT_30STG if r[0] == vessel_type]
    return pd.DataFrame(rows, columns=['vessel_type','block_no','관철','덕트','전장','도장','배선'])


def default_50stg_df(vessel_type: str) -> pd.DataFrame:
    rows = [r for r in _INIT_50STG if r[0] == vessel_type]
    return pd.DataFrame(rows, columns=[
        'vessel_type','deck',
        '선각취부','선각용접','선각FLOOR곡직','선각WALL곡직',
        '목의화기1차','도장PB','도장TU','배선','보온',
        '목의화기2차','피복','BP검사','판넬설치','결선','가구','장판','스커트'])


def default_inout_df() -> pd.DataFrame:
    return pd.DataFrame(_INIT_INOUT, columns=['area'] + _IO_WORK_COLS)


def default_sequence_df(stg: str = None) -> pd.DataFrame:
    rows = [r for r in _INIT_SEQUENCE if (stg is None or r[0] == stg)]
    return pd.DataFrame(rows, columns=[
        'stg','seq_no','work_code','work_name','duration_type',
        'fixed_days','exception_blocks','cal_type','ref_col'])


def default_deck_order_df(vessel_type: str = None) -> pd.DataFrame:
    rows = [r for r in _INIT_DECK_ORDER if (vessel_type is None or r[0] == vessel_type)]
    return pd.DataFrame(rows, columns=['vessel_type','deck','order_no'])


def default_rules_df() -> pd.DataFrame:
    return pd.DataFrame(_INIT_RULES, columns=[
        'rule_id','vessel_type','rule_name','pre_area','pre_work',
        'pre_event','offset_days','suc_area','suc_work','is_active'])


# ═══════════════════════════════════════════════════════════════
# 앵커 이벤트 Excel 양식 생성
# ═══════════════════════════════════════════════════════════════

def generate_anchor_template(vessel_type: str = "LNG") -> bytes:
    """앵커 이벤트 입력용 Excel 양식 바이트 반환 (openpyxl)
    vessel_type: "LNG" 또는 "CONT" — 해당 선종의 블럭/데크를 미리 채워 반환
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl 미설치")
        return b""

    wb = Workbook()

    # ── 공통 스타일 ──────────────────────────────────────────
    HDR_FILL     = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT     = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    BODY_FONT    = Font(name="맑은 고딕", size=10)
    CENTER       = Alignment(horizontal="center", vertical="center")
    THIN         = Side(style="thin", color="AAAAAA")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    DATE_FMT     = "YYYY-MM-DD"
    PREFILL_FILL = PatternFill("solid", fgColor="DEEAF1")  # 연파랑 — 미리채운 셀

    def _write_header(ws, headers, col_widths):
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = CENTER; cell.border = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22

    def _set_cell(ws, r, c, value, fill=None, date_fmt=False):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = BODY_FONT; cell.alignment = CENTER; cell.border = BORDER
        if fill:
            cell.fill = fill
        if date_fmt:
            cell.number_format = DATE_FMT

    def _blank_date(ws, r, c):
        cell = ws.cell(row=r, column=c, value="")
        cell.font = BODY_FONT; cell.alignment = CENTER; cell.border = BORDER
        cell.number_format = DATE_FMT

    # ── 블럭/데크 목록 (선종 기준, 데크 순서대로) ────────────
    dk_order = {r[1]: r[2] for r in _INIT_DECK_ORDER if r[0] == vessel_type}
    blocks_sorted = sorted(
        [(r[1], r[2]) for r in _INIT_BLOCK_DECK_MAP if r[0] == vessel_type],
        key=lambda x: (dk_order.get(x[1], 99), x[0])
    )
    decks_sorted = sorted(
        [r[1] for r in _INIT_DECK_ORDER if r[0] == vessel_type],
        key=lambda d: dk_order.get(d, 99)
    )

    # ── Sheet 1: 호선정보 ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "호선정보"
    ws1.freeze_panes = "A2"
    _write_header(ws1,
                  ["호선번호", "선종(LNG/CONT)", "거주구탑재예정일", "비고"],
                  [14, 16, 20, 30])
    for r in range(2, 22):
        _set_cell(ws1, r, 1, "")
        _set_cell(ws1, r, 2, "")   # 선종은 직접 입력 (LNG 또는 CONT)
        _blank_date(ws1, r, 3)
        _set_cell(ws1, r, 4, "")

    # ── Sheet 2: 30STG_앵커 ──────────────────────────────────
    ws2 = wb.create_sheet("30STG_앵커")
    ws2.freeze_panes = "D2"
    _write_header(ws2,
                  ["호선번호", "선종(LNG/CONT)", "블럭번호", "블럭입고일", "블럭탑재일"],
                  [14, 16, 14, 16, 16])
    for ri, (blk, _dk) in enumerate(blocks_sorted, 2):
        _set_cell(ws2, ri, 1, "")
        _set_cell(ws2, ri, 2, vessel_type, fill=PREFILL_FILL)
        _set_cell(ws2, ri, 3, blk,         fill=PREFILL_FILL)
        _blank_date(ws2, ri, 4)
        _blank_date(ws2, ri, 5)

    # ── Sheet 3: 50STG_앵커 ──────────────────────────────────
    ws3 = wb.create_sheet("50STG_앵커")
    ws3.freeze_panes = "D2"
    _write_header(ws3,
                  ["호선번호", "선종(LNG/CONT)", "데크",
                   "블럭 탑재 종료일",
                   "선각검사 완료일"],
                  [14, 16, 12, 18, 18])
    for ri, dk in enumerate(decks_sorted, 2):
        _set_cell(ws3, ri, 1, "")
        _set_cell(ws3, ri, 2, vessel_type, fill=PREFILL_FILL)
        _set_cell(ws3, ri, 3, dk,          fill=PREFILL_FILL)
        _blank_date(ws3, ri, 4)
        _blank_date(ws3, ri, 5)

    # ── 안내 시트 ─────────────────────────────────────────────
    ws_guide = wb.create_sheet("입력안내", 0)
    guide_rows = [
        ["", ""],
        [f"  엑셀 회의록 대체 — 앵커 이벤트 입력 양식 ({vessel_type})", ""],
        ["", ""],
        ["  시트명",          "내용"],
        ["  호선정보",        "진행 중인 호선 목록 및 거주구탑재 예정일"],
        ["  30STG_앵커",      f"블럭별 입고일·탑재일 ({vessel_type} · {len(blocks_sorted)}개 블럭 미리 채워짐)"],
        ["  50STG_앵커",      f"데크별 선각공정 완료일 ({vessel_type} · {len(decks_sorted)}개 데크 미리 채워짐)"],
        ["", ""],
        ["  공통 주의사항",   ""],
        ["  ·", "날짜 형식: YYYY-MM-DD (예: 2026-01-15)"],
        ["  ·", "선종: LNG 또는 CONT 정확히 입력"],
        ["  ·", "연파랑 셀(블럭번호·데크)은 미리 채워진 값 — 수정 금지 / 선종(호선정보)은 직접 입력"],
        ["  ·", "업로드 후 기존 동일 키(호선+블럭/데크)는 덮어씀"],
        ["", ""],
        ["  50STG 앵커 컬럼 설명", ""],
        ["  블럭 탑재 종료일", "마지막 블럭 탑재 종료일 (취부·용접·곡직은 기준정보 소요일로 자동 계산)"],
        ["  선각검사 완료일",  "선각검사 완료일 (50STG 의장공정 시작 기준점) — 비워두면 자동 계산"],
    ]
    for r_idx, (a, b) in enumerate(guide_rows, 1):
        ca = ws_guide.cell(row=r_idx, column=1, value=a)
        cb = ws_guide.cell(row=r_idx, column=2, value=b)
        if r_idx == 2:
            ca.font = Font(name="맑은 고딕", bold=True, size=13, color="1F4E79")
        elif r_idx == 4:
            ca.font = Font(name="맑은 고딕", bold=True, size=10)
            cb.font = Font(name="맑은 고딕", bold=True, size=10)
        else:
            ca.font = Font(name="맑은 고딕", size=10)
            cb.font = Font(name="맑은 고딕", size=10)
    ws_guide.column_dimensions["A"].width = 20
    ws_guide.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
# 앵커 이벤트 조회
# ═══════════════════════════════════════════════════════════════

@st.cache_data(ttl=60)
def get_vessels() -> pd.DataFrame:
    return run_query(
        "SELECT vessel_no,vessel_type,거주구탑재예정일,비고 FROM lq_meet_vessel ORDER BY vessel_no"
    )


@st.cache_data(ttl=60)
def get_anchor_30stg(vessel_no: str = None) -> pd.DataFrame:
    sql = "SELECT vessel_no,block_no,blk_in_date,blk_out_date FROM lq_meet_anchor_30stg"
    if vessel_no:
        return run_query(sql + " WHERE vessel_no=? ORDER BY block_no", params=(vessel_no,))
    return run_query(sql + " ORDER BY vessel_no,block_no")


@st.cache_data(ttl=60)
def get_anchor_50stg(vessel_no: str = None) -> pd.DataFrame:
    sql = """SELECT vessel_no,deck,mount_start_date,
                    sunggak_attach_end,sunggak_weld_end,
                    floor_straight_date,wall_straight_date,inspection_date
             FROM lq_meet_anchor_50stg"""
    if vessel_no:
        return run_query(sql + " WHERE vessel_no=? ORDER BY deck", params=(vessel_no,))
    return run_query(sql + " ORDER BY vessel_no,deck")


# ═══════════════════════════════════════════════════════════════
# 앵커 이벤트 Excel 업로드 → DB 저장
# ═══════════════════════════════════════════════════════════════

def _to_date_str(val) -> str | None:
    """openpyxl 셀 값 → 'YYYY-MM-DD' 문자열"""
    if val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return str(val)[:10]
    s = str(val).strip()
    return None if s in ("", "None") else s


def _calc_sunggak_from_ref(mount_end: datetime.date, deck: str, vessel_type: str) -> dict:
    """탑재 종료일 + 기준정보 소요일 → 선각 중간 날짜 자동 계산"""
    try:
        cal_df = get_calendar()
        holidays = {str(d)[:10] for d in cal_df["cal_date"]} if not cal_df.empty else set()
    except Exception:
        holidays = set()

    def _nwd(d):
        while d.weekday() == 6 or str(d) in holidays:
            d += datetime.timedelta(days=1)
        return d

    def _awd(start, n):
        if n <= 0:
            return start
        rem = n - 1
        cur = start
        while rem > 0:
            cur += datetime.timedelta(days=1)
            if cur.weekday() != 6 and str(cur) not in holidays:
                rem -= 1
        return cur

    result = {
        "sunggak_attach_end": None,
        "sunggak_weld_end": None,
        "floor_straight_date": None,
        "wall_straight_date": None,
    }
    try:
        ref_df = get_ref_50stg(vessel_type)
        if ref_df.empty:
            return result
        ref_row = ref_df[ref_df["deck"] == deck]
        if ref_row.empty:
            return result
        ref = ref_row.iloc[0]
        prev = mount_end
        for col_key, db_col in [
            ("선각취부",      "sunggak_attach_end"),
            ("선각용접",      "sunggak_weld_end"),
            ("선각FLOOR곡직", "floor_straight_date"),
            ("선각WALL곡직",  "wall_straight_date"),
        ]:
            days = int(ref.get(col_key, 0) or 0)
            if days == 0:
                continue
            start = _nwd(prev + datetime.timedelta(days=1))
            end   = _awd(start, days)
            result[db_col] = end
            prev = end
    except Exception as e:
        logger.error("선각 날짜 계산 오류: %s", e)
    return result


def upload_anchor_excel(file_bytes: bytes) -> dict:
    """앵커 이벤트 Excel 양식 파싱 후 DB Upsert
    반환: {"vessel": int, "anchor30": int, "anchor50": int, "errors": list}
    """
    from openpyxl import load_workbook

    result = {"vessel": 0, "anchor30": 0, "anchor50": 0, "errors": []}

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        msg = str(e)
        if "zip file" in msg.lower() or "not a zip" in msg.lower():
            result["errors"].append(
                "파일 형식 오류: .xls(구버전) 파일은 지원되지 않습니다. "
                "Excel에서 [다른 이름으로 저장 → Excel 통합 문서(.xlsx)]로 저장 후 다시 업로드하세요."
            )
        else:
            result["errors"].append(f"파일 읽기 오류: {msg}")
        return result

    # ── Sheet: 호선정보 ────────────────────────────────────────
    if "호선정보" in wb.sheetnames:
        ws = wb["호선정보"]
        conn = get_connection()
        cur  = conn.cursor()
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = (list(row) + [None] * 4)[:4]
            vessel_no, vessel_type, 거주구일, 비고 = vals
            if not vessel_no or str(vessel_no).strip() == "":
                continue
            vessel_no   = str(vessel_no).strip()
            vessel_type = str(vessel_type).strip() if vessel_type else ""
            if vessel_type not in ("LNG", "CONT"):
                result["errors"].append(f"호선 {vessel_no}: 선종이 LNG/CONT가 아닙니다.")
                continue
            d = _to_date_str(거주구일)
            cur.execute("""
                IF EXISTS (SELECT 1 FROM lq_meet_vessel WHERE vessel_no=?)
                    UPDATE lq_meet_vessel
                    SET vessel_type=?,거주구탑재예정일=?,비고=?,updated_at=GETDATE()
                    WHERE vessel_no=?
                ELSE
                    INSERT INTO lq_meet_vessel (vessel_no,vessel_type,거주구탑재예정일,비고)
                    VALUES (?,?,?,?)
            """, (vessel_no, vessel_type, d, 비고, vessel_no,
                  vessel_no, vessel_type, d, 비고))
            result["vessel"] += 1
        conn.commit()
        conn.close()
        get_vessels.clear()

    # ── Sheet: 30STG_앵커 ─────────────────────────────────────
    if "30STG_앵커" in wb.sheetnames:
        ws = wb["30STG_앵커"]
        conn = get_connection()
        cur  = conn.cursor()
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = (list(row) + [None] * 5)[:5]
            vessel_no, _, block_no, blk_in, blk_out = vals
            if not vessel_no or not block_no:
                continue
            vno = str(vessel_no).strip()
            bno = str(block_no).strip()
            d_in  = _to_date_str(blk_in)
            d_out = _to_date_str(blk_out)
            cur.execute("""
                IF EXISTS (SELECT 1 FROM lq_meet_anchor_30stg WHERE vessel_no=? AND block_no=?)
                    UPDATE lq_meet_anchor_30stg
                    SET blk_in_date=?,blk_out_date=?,updated_at=GETDATE()
                    WHERE vessel_no=? AND block_no=?
                ELSE
                    INSERT INTO lq_meet_anchor_30stg (vessel_no,block_no,blk_in_date,blk_out_date)
                    VALUES (?,?,?,?)
            """, (vno, bno, d_in, d_out, vno, bno,
                  vno, bno, d_in, d_out))
            result["anchor30"] += 1
        conn.commit()
        conn.close()
        get_anchor_30stg.clear()

    # ── Sheet: 50STG_앵커 ─────────────────────────────────────
    if "50STG_앵커" in wb.sheetnames:
        ws = wb["50STG_앵커"]
        conn = get_connection()
        cur  = conn.cursor()
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = (list(row) + [None] * 5)[:5]
            vessel_no, vessel_type_cell, deck, mount_end_raw, insp_end_raw = vals
            if not vessel_no or not deck:
                continue
            vno   = str(vessel_no).strip()
            dk    = str(deck).strip()
            vtype = str(vessel_type_cell).strip() if vessel_type_cell else ""
            d_mount = _to_date_str(mount_end_raw)
            d_insp  = _to_date_str(insp_end_raw)

            # 취부·용접·곡직 날짜 — 기준정보 소요일로 자동 계산
            d_attach = d_weld = d_floor = d_wall = None
            if d_mount and vtype in ("LNG", "CONT"):
                try:
                    mount_dt = datetime.date.fromisoformat(d_mount)
                    calc = _calc_sunggak_from_ref(mount_dt, dk, vtype)
                    d_attach = str(calc["sunggak_attach_end"]) if calc.get("sunggak_attach_end") else None
                    d_weld   = str(calc["sunggak_weld_end"])   if calc.get("sunggak_weld_end")   else None
                    d_floor  = str(calc["floor_straight_date"]) if calc.get("floor_straight_date") else None
                    d_wall   = str(calc["wall_straight_date"])  if calc.get("wall_straight_date")  else None
                    # 선각검사일 미입력 시 wall 다음 영업일로 자동 계산
                    if not d_insp and d_wall:
                        wall_dt = datetime.date.fromisoformat(d_wall)
                        d_insp = str(wall_dt + datetime.timedelta(days=1))
                except Exception as ce:
                    logger.error("50STG 날짜 계산 오류 (호선=%s, 데크=%s): %s", vno, dk, ce)

            cur.execute("""
                IF EXISTS (SELECT 1 FROM lq_meet_anchor_50stg WHERE vessel_no=? AND deck=?)
                    UPDATE lq_meet_anchor_50stg
                    SET mount_start_date=?,sunggak_attach_end=?,sunggak_weld_end=?,
                        floor_straight_date=?,wall_straight_date=?,inspection_date=?,
                        updated_at=GETDATE()
                    WHERE vessel_no=? AND deck=?
                ELSE
                    INSERT INTO lq_meet_anchor_50stg
                      (vessel_no,deck,mount_start_date,sunggak_attach_end,sunggak_weld_end,
                       floor_straight_date,wall_straight_date,inspection_date)
                    VALUES (?,?,?,?,?,?,?,?)
            """, (vno, dk, d_mount, d_attach, d_weld, d_floor, d_wall, d_insp, vno, dk,
                  vno, dk, d_mount, d_attach, d_weld, d_floor, d_wall, d_insp))
            result["anchor50"] += 1
        conn.commit()
        conn.close()
        get_anchor_50stg.clear()

    return result
